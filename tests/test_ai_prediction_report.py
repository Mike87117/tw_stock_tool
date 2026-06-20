import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

import ai_prediction_report


def _detail_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Window": 1,
                "Train Start": pd.Timestamp("2024-01-01"),
                "Train End": pd.Timestamp("2024-01-08"),
                "Test Start": pd.Timestamp("2024-01-09"),
                "Test End": pd.Timestamp("2024-01-12"),
                "Train Rows": 8,
                "Test Rows": 4,
                "Feature Count": 4,
                "Target Column": "Target_Up_5D",
                "Accuracy": 0.75,
                "Precision": 0.8,
                "Recall": 0.67,
                "F1": 0.73,
                "Train Positive Rate %": 50.0,
                "Test Positive Rate %": 50.0,
                "Predicted Positive Rate %": 60.0,
                "Error": "",
            },
            {
                "Window": 2,
                "Train Start": pd.Timestamp("2024-01-05"),
                "Train End": pd.Timestamp("2024-01-12"),
                "Test Start": pd.Timestamp("2024-01-13"),
                "Test End": pd.Timestamp("2024-01-16"),
                "Train Rows": 8,
                "Test Rows": 4,
                "Feature Count": 4,
                "Target Column": "Target_Up_5D",
                "Accuracy": None,
                "Precision": None,
                "Recall": None,
                "F1": None,
                "Train Positive Rate %": None,
                "Test Positive Rate %": None,
                "Predicted Positive Rate %": None,
                "Error": "model failed",
            },
        ],
        columns=ai_prediction_report.BASELINE_RESULT_COLUMNS,
    )


class AIPredictionReportTest(unittest.TestCase):
    def test_run_ai_prediction_report_calls_baseline_model(self) -> None:
        detail = _detail_df()
        with patch.object(
            ai_prediction_report,
            "run_baseline_ml_model",
            return_value=detail,
        ) as mocked:
            frames = ai_prediction_report.run_ai_prediction_report(
                "2330",
                period="5y",
                horizon=5,
                train_size=8,
                test_size=4,
                step_size=4,
                force_refresh=True,
                dropna=False,
                n_estimators=5,
                random_state=7,
            )

        mocked.assert_called_once_with(
            stock_id="2330",
            period="5y",
            horizon=5,
            train_size=8,
            test_size=4,
            step_size=4,
            force_refresh=True,
            dropna=False,
            n_estimators=5,
            random_state=7,
        )
        self.assertIn("Summary", frames)
        self.assertIn("Detail", frames)
        self.assertIn("Errors", frames)

    def test_summary_contains_expected_values(self) -> None:
        frames = ai_prediction_report.build_report_frames(
            _detail_df(),
            stock_id="2330",
            period="5y",
            horizon=5,
            train_size=8,
            test_size=4,
            step_size=4,
        )
        summary = frames["Summary"].iloc[0]

        self.assertEqual(summary["Stock"], "2330")
        self.assertEqual(summary["Windows"], 2)
        self.assertEqual(summary["Error Windows"], 1)
        self.assertEqual(summary["Avg Accuracy"], 0.75)

    def test_errors_frame_contains_only_failed_rows(self) -> None:
        frames = ai_prediction_report.build_report_frames(
            _detail_df(),
            stock_id="2330",
            period="5y",
            horizon=5,
            train_size=8,
            test_size=4,
            step_size=4,
        )

        self.assertEqual(len(frames["Errors"]), 1)
        self.assertEqual(frames["Errors"].iloc[0]["Error"], "model failed")

    def test_export_excel_creates_required_sheets(self) -> None:
        frames = ai_prediction_report.build_report_frames(
            _detail_df(),
            stock_id="2330",
            period="5y",
            horizon=5,
            train_size=8,
            test_size=4,
            step_size=4,
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "ai_prediction_report.xlsx"
            result = ai_prediction_report.export_ai_prediction_report_excel(
                frames,
                stock_id="2330",
                output=str(output_path),
            )
            workbook = load_workbook(result, read_only=True)

            self.assertEqual(result, output_path)
            self.assertIn("Summary", workbook.sheetnames)
            self.assertIn("Detail", workbook.sheetnames)
            self.assertIn("Errors", workbook.sheetnames)
            workbook.close()

    def test_export_excel_default_path(self) -> None:
        frames = ai_prediction_report.build_report_frames(
            _detail_df(),
            stock_id="2330",
            period="5y",
            horizon=5,
            train_size=8,
            test_size=4,
            step_size=4,
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            with patch.object(ai_prediction_report, "OUTPUT_DIR", Path(tmp_dir)):
                result = ai_prediction_report.export_ai_prediction_report_excel(
                    frames,
                    stock_id="2330",
                    output="",
                )

            self.assertEqual(result, Path(tmp_dir) / "2330_ai_prediction_report.xlsx")
            self.assertTrue(result.exists())

    def test_parse_args(self) -> None:
        args = ai_prediction_report._parse_args(
            [
                "--stock",
                "2330",
                "--period",
                "5y",
                "--horizon",
                "5",
                "--train-size",
                "8",
                "--test-size",
                "4",
                "--step-size",
                "4",
                "--n-estimators",
                "5",
                "--random-state",
                "7",
                "--no-dropna",
                "--output",
            ]
        )

        self.assertEqual(args.stock, "2330")
        self.assertEqual(args.period, "5y")
        self.assertEqual(args.horizon, 5)
        self.assertEqual(args.train_size, 8)
        self.assertEqual(args.test_size, 4)
        self.assertEqual(args.step_size, 4)
        self.assertEqual(args.n_estimators, 5)
        self.assertEqual(args.random_state, 7)
        self.assertFalse(args.dropna)
        self.assertEqual(args.output, "")


if __name__ == "__main__":
    unittest.main()

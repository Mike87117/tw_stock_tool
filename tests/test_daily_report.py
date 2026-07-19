import tempfile
import sys
import unittest
from types import SimpleNamespace
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from tw_stock_tool.reports import daily_report


def _ranking_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Rank": 1,
                "Stock": "2330",
                "Signal": "BUY",
                "Score": 6.0,
                "Close": 900.0,
                "Volume_Ratio": 1.8,
                "RSI": 62.0,
                "Analysis": "strong",
                "Status": "OK",
                "Error": "",
            },
            {
                "Rank": 2,
                "Stock": "2454",
                "Signal": "BUY",
                "Score": 6.0,
                "Close": 1200.0,
                "Volume_Ratio": 2.5,
                "RSI": 58.0,
                "Analysis": "volume",
                "Status": "OK",
                "Error": "",
            },
            {
                "Rank": 3,
                "Stock": "2317",
                "Signal": "WATCH",
                "Score": 5.0,
                "Close": 180.0,
                "Volume_Ratio": 1.2,
                "RSI": 54.0,
                "Analysis": "watch",
                "Status": "OK",
                "Error": "",
            },
            {
                "Rank": 4,
                "Stock": "2308",
                "Signal": "HOLD",
                "Score": 8.0,
                "Close": 300.0,
                "Volume_Ratio": 4.0,
                "RSI": 70.0,
                "Analysis": "hold",
                "Status": "OK",
                "Error": "",
            },
            {
                "Rank": None,
                "Stock": "9999",
                "Signal": "",
                "Score": None,
                "Close": None,
                "Volume_Ratio": None,
                "RSI": None,
                "Analysis": "",
                "Status": "ERROR",
                "Error": "bad stock",
            },
        ]
    )


class DailyReportTest(unittest.TestCase):
    def test_build_summary(self) -> None:
        ranking = _ranking_df()
        candidates = daily_report.filter_candidates(ranking, min_score=4, top=20)

        summary = daily_report.build_summary(ranking, candidates, report_date="2026-06-19")

        self.assertEqual(summary.loc[0, "Report Date"], "2026-06-19")
        self.assertEqual(summary.loc[0, "Stocks Scanned"], 5)
        self.assertEqual(summary.loc[0, "Candidates"], 3)
        self.assertEqual(summary.loc[0, "BUY Count"], 2)
        self.assertEqual(summary.loc[0, "WATCH Count"], 1)
        self.assertEqual(summary.loc[0, "Average Score"], 5.67)

    def test_candidate_filter(self) -> None:
        result = daily_report.filter_candidates(
            _ranking_df(),
            signals=("BUY", "WATCH"),
            min_score=5,
            top=20,
        )

        self.assertEqual(result["Stock"].tolist(), ["2454", "2330", "2317"])
        self.assertNotIn("2308", result["Stock"].tolist())
        self.assertEqual(result["Rank"].tolist(), [1, 2, 3])

    def test_sorting_by_score_then_volume_ratio(self) -> None:
        result = daily_report.filter_candidates(_ranking_df(), min_score=4, top=2)

        self.assertEqual(result["Stock"].tolist(), ["2454", "2330"])

    def test_error_sheet_contains_failed_rows(self) -> None:
        ranking = _ranking_df()
        candidates = daily_report.filter_candidates(ranking)
        summary = daily_report.build_summary(ranking, candidates, report_date="2026-06-19")

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "daily_report.xlsx"
            output = daily_report.export_daily_report(summary, candidates, ranking, str(path))
            workbook = load_workbook(output, read_only=True)
            rows = list(workbook["Errors"].iter_rows(values_only=True))
            workbook.close()

        self.assertGreaterEqual(len(rows), 2)
        self.assertIn("9999", rows[1])

    def test_export_excel_creates_expected_sheets(self) -> None:
        ranking = _ranking_df()
        candidates = daily_report.filter_candidates(ranking)
        summary = daily_report.build_summary(ranking, candidates, report_date="2026-06-19")

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "daily_report.xlsx"
            output = daily_report.export_daily_report(summary, candidates, ranking, str(path))
            workbook = load_workbook(output, read_only=True)
            sheetnames = workbook.sheetnames
            workbook.close()

        self.assertIn("Summary", sheetnames)
        self.assertIn("Candidates", sheetnames)
        self.assertIn("All", sheetnames)
        self.assertIn("Errors", sheetnames)

    def test_collect_stock_ids_auto_stock_list_calls_updater_and_has_priority(self) -> None:
        updater_df = pd.DataFrame([{"Stock": "2330"}, {"Stock": "2454"}])
        with patch.object(
            daily_report.stock_list_updater_module,
            "update_stock_list",
            return_value=(updater_df, "stocks.txt"),
        ) as updater_mock:
            result = daily_report.collect_stock_ids(
                ["9999"],
                "ignored.txt",
                auto_stock_list=True,
                stock_market="twse",
                stock_list_output="stocks.txt",
                allow_partial_stock_list=True,
            )

        updater_mock.assert_called_once_with(
            market="twse",
            output="stocks.txt",
            allow_partial=True,
        )
        self.assertEqual(result, ["2330", "2454"])

    def test_collect_stock_ids_applies_stock_limit(self) -> None:
        result = daily_report.collect_stock_ids(
            ["2330", "2317", "2454"],
            None,
            stock_limit=2,
        )

        self.assertEqual(result, ["2330", "2317"])

    def test_collect_stock_ids_applies_stock_sample(self) -> None:
        first = daily_report.collect_stock_ids(
            ["2330", "2317", "2454", "2308"],
            None,
            stock_sample=2,
            random_state=7,
        )
        second = daily_report.collect_stock_ids(
            ["2330", "2317", "2454", "2308"],
            None,
            stock_sample=2,
            random_state=7,
        )

        self.assertEqual(first, second)
        self.assertEqual(len(first), 2)

    def test_data_limitations_from_ranking_with_failures(self) -> None:
        ranking = _ranking_df()
        limitations = daily_report.build_data_limitations_from_ranking(ranking)
        self.assertEqual(len(limitations), 1)
        self.assertEqual(limitations[0], "9999: ERROR - bad stock")

    def test_data_limitations_from_ranking_all_ok(self) -> None:
        ranking = pd.DataFrame([
            {"Stock": "2330", "Status": "OK", "Error": ""},
            {"Stock": "2317", "Status": "OK", "Error": ""},
        ])
        limitations = daily_report.build_data_limitations_from_ranking(ranking)
        self.assertEqual(limitations, [])

    def test_data_limitations_from_ranking_empty(self) -> None:
        limitations = daily_report.build_data_limitations_from_ranking(pd.DataFrame())
        self.assertEqual(limitations, [])
        limitations_none = daily_report.build_data_limitations_from_ranking(None)
        self.assertEqual(limitations_none, [])

    def test_data_limitations_from_ranking_truncation(self) -> None:
        rows = [{"Stock": str(i), "Status": "ERROR", "Error": "fail"} for i in range(1, 15)]
        ranking = pd.DataFrame(rows)
        limitations = daily_report.build_data_limitations_from_ranking(ranking, max_items=10)
        self.assertEqual(len(limitations), 11)
        self.assertEqual(limitations[0], "1: ERROR - fail")
        self.assertEqual(limitations[-1], "... and 4 more failed stock(s).")



class CandidateBacktestValidationTest(unittest.TestCase):
    def _candidates(self) -> pd.DataFrame:
        return pd.DataFrame([
            {"Rank": 2, "Stock": "2454", "Signal": "BUY", "Score": 6.2},
            {"Rank": 1, "Stock": "2330", "Signal": "WATCH", "Score": 5.8},
            {"Rank": 3, "Stock": "2317", "Signal": "BUY", "Score": 5.1},
        ])

    def _analysis(self) -> SimpleNamespace:
        frame = pd.DataFrame(
            {
                "Open": [10.0, 11.0],
                "Close": [10.0, 12.0],
                "Signal": ["HOLD", "BUY"],
            },
            index=pd.date_range("2026-01-01", periods=2),
        )
        return SimpleNamespace(indicator_df=frame, signal_df=frame)

    def _result(self) -> SimpleNamespace:
        return SimpleNamespace(
            start_date=pd.Timestamp("2026-01-01"),
            end_date=pd.Timestamp("2026-01-02"),
            total_return_pct=12.345,
            buy_hold_return_pct=20.125,
            trade_count=2,
            win_rate_pct=50.555,
            max_drawdown_pct=-3.333,
            sharpe_ratio=1.234,
        )

    def test_validation_is_opt_in_and_preserves_candidate_order(self) -> None:
        with patch.object(daily_report, "analyze_stock") as analyze, patch.object(
            daily_report, "run_backtest_result", return_value=self._result()
        ) as run_backtest, patch.dict(
            daily_report.STRATEGIES, {"ma_cross_strategy": lambda frame: frame}, clear=False
        ):
            analyze.side_effect = [self._analysis(), self._analysis()]
            highlights, limitations = daily_report.run_candidate_backtest_validation(
                self._candidates(),
                validate_top=2,
                strategy="ma_cross",
                period="2y",
                interval="1wk",
                auto_adjust=True,
                force_refresh=True,
                initial_capital=200000,
                fee_rate=0.001,
                tax_rate=0.002,
                position_size=0.5,
            )

        self.assertEqual(highlights["Stock"].tolist(), ["2454", "2330"])
        self.assertEqual(highlights["Status"].tolist(), ["OK", "OK"])
        self.assertEqual(highlights["Total Return %"].tolist(), [12.35, 12.35])
        self.assertEqual(highlights.columns.tolist(), daily_report.BACKTEST_HIGHLIGHT_COLUMNS)
        self.assertEqual(limitations, [])
        self.assertEqual(analyze.call_args_list[0].kwargs["interval"], "1wk")
        self.assertEqual(run_backtest.call_args.kwargs["position_size"], 0.5)
        self.assertEqual(run_backtest.call_args.kwargs["interval"], "1wk")

    def test_validation_failure_isolated_and_zero_is_noop(self) -> None:
        with patch.object(daily_report, "analyze_stock") as analyze, patch.object(
            daily_report, "run_backtest_result", return_value=self._result()
        ) as run_backtest, patch.dict(
            daily_report.STRATEGIES, {"score_strategy": lambda frame: frame}, clear=False
        ):
            empty, limits = daily_report.run_candidate_backtest_validation(
                self._candidates(),
                validate_top=0,
                strategy="score",
                period="1y",
                interval="1d",
                auto_adjust=False,
                force_refresh=False,
                initial_capital=100000,
                fee_rate=0.001425,
                tax_rate=0.003,
                position_size=1.0,
            )
            self.assertTrue(empty.empty)
            self.assertEqual(limits, [])
            analyze.assert_not_called()
            run_backtest.assert_not_called()

            analyze.reset_mock()
            analyze.side_effect = [ValueError("first\nerror"), self._analysis()]
            highlights, limits = daily_report.run_candidate_backtest_validation(
                self._candidates(),
                validate_top=2,
                strategy="score",
                period="1y",
                interval="1d",
                auto_adjust=False,
                force_refresh=False,
                initial_capital=100000,
                fee_rate=0.001425,
                tax_rate=0.003,
                position_size=1.0,
            )

        self.assertEqual(highlights["Status"].tolist(), ["ERROR", "OK"])
        self.assertTrue(pd.isna(highlights.iloc[0]["Total Return %"]))
        self.assertEqual(limits, ["Backtest validation for 2454 failed: first error"])

    def test_markdown_renders_scalar_backtest_highlights(self) -> None:
        report = daily_report.build_daily_report_data(
            report_date="2026-07-19",
            stock_universe=["2330"],
            screening_results=pd.DataFrame([{"Stocks Scanned": 1, "Candidates": 1, "BUY Count": 1, "WATCH Count": 0}]),
            backtest_highlights=pd.DataFrame([{
                "Rank": 1, "Stock": "2330", "Signal": "BUY", "Score": 6.0,
                "Strategy": "ma_cross", "Status": "OK", "Start Date": "2026-01-01",
                "End Date": "2026-07-19", "Total Return %": 12.35,
                "Buy and Hold Return %": 20.13, "Trade Count": 2, "Win Rate %": 50.0,
                "Max Drawdown %": -3.33, "Sharpe Ratio": 1.23, "Error": "",
            }]),
        )
        markdown = daily_report.render_daily_report_markdown(report)
        self.assertIn("## Backtest Highlights", markdown)
        self.assertIn("12.35", markdown)
        self.assertNotIn("Trades", markdown)
        self.assertNotIn("Equity Curve", markdown)

class CandidateValidationBoundaryTest(unittest.TestCase):
    def _candidates(self, count: int = 3) -> pd.DataFrame:
        rows = [
            {"Rank": 2, "Stock": "2454", "Signal": "BUY", "Score": 6.2},
            {"Rank": 1, "Stock": "2330", "Signal": "WATCH", "Score": 5.8},
            {"Rank": 3, "Stock": "2317", "Signal": "BUY", "Score": 5.1},
        ]
        return pd.DataFrame(rows[:count])

    def _frame(self, marker: str) -> pd.DataFrame:
        return pd.DataFrame(
            {
                "Open": [10.0, 11.0],
                "Close": [10.0, 12.0],
                "Signal": ["HOLD", "BUY"],
                "Marker": [marker, marker],
            },
            index=pd.date_range("2026-01-01", periods=2),
        )

    def _analysis(self) -> SimpleNamespace:
        return SimpleNamespace(
            indicator_df=self._frame("indicator"),
            signal_df=self._frame("signal"),
        )

    def _result(self) -> SimpleNamespace:
        return SimpleNamespace(
            start_date=pd.Timestamp("2026-01-01"),
            end_date=pd.Timestamp("2026-01-02"),
            total_return_pct=12.345,
            buy_hold_return_pct=20.125,
            trade_count=2,
            win_rate_pct=50.555,
            max_drawdown_pct=-3.333,
            sharpe_ratio=1.234,
        )

    def _run(self, strategy: str, candidates: pd.DataFrame | None = None):
        analysis = self._analysis()
        received = []

        def strategy_func(frame):
            received.append(frame)
            return frame

        with patch.object(daily_report, "analyze_stock", return_value=analysis) as analyze, patch.object(
            daily_report, "run_backtest_result", return_value=self._result()
        ) as run_backtest, patch.dict(
            daily_report.STRATEGIES,
            {f"{strategy}_strategy": strategy_func},
            clear=False,
        ):
            highlights, limitations = daily_report.run_candidate_backtest_validation(
                candidates if candidates is not None else self._candidates(1),
                validate_top=1,
                strategy=strategy,
                period="2y",
                interval="1wk",
                auto_adjust=True,
                force_refresh=True,
                initial_capital=200000,
                fee_rate=0.001,
                tax_rate=0.002,
                position_size=0.5,
            )
        return analysis, received, analyze, run_backtest, highlights, limitations

    def test_score_uses_signal_and_non_score_strategies_use_indicator(self) -> None:
        analysis, received, analyze, run_backtest, _, _ = self._run("score")
        self.assertIs(received[0], analysis.signal_df)
        self.assertEqual(analyze.call_args.kwargs, {
            "stock_id": "2454", "period": "2y", "interval": "1wk",
            "auto_adjust": True, "force_refresh": True,
        })
        self.assertEqual(run_backtest.call_args.kwargs, {
            "initial_capital": 200000, "fee_rate": 0.001,
            "tax_rate": 0.002, "position_size": 0.5, "interval": "1wk",
        })

        for strategy in ("ma_cross", "rsi"):
            analysis, received, _, _, _, _ = self._run(strategy)
            self.assertIs(received[0], analysis.indicator_df)

    def test_fewer_candidates_empty_schema_and_unsupported_strategy(self) -> None:
        analysis = self._analysis()
        with patch.object(daily_report, "analyze_stock", return_value=analysis) as analyze, patch.object(
            daily_report, "run_backtest_result", return_value=self._result()
        ) as run_backtest, patch.dict(
            daily_report.STRATEGIES, {"ma_cross_strategy": lambda frame: frame}, clear=False
        ):
            highlights, limitations = daily_report.run_candidate_backtest_validation(
                self._candidates(2), validate_top=5, strategy="ma_cross", period="1y",
                interval="1d", auto_adjust=False, force_refresh=False,
                initial_capital=100000, fee_rate=0.001425, tax_rate=0.003, position_size=1.0,
            )
            self.assertEqual(highlights["Stock"].tolist(), ["2454", "2330"])
            self.assertEqual(len(highlights), 2)
            self.assertEqual(limitations, [])
            self.assertEqual(analyze.call_count, 2)
            self.assertEqual(run_backtest.call_count, 2)

            analyze.reset_mock()
            run_backtest.reset_mock()
            empty, empty_limits = daily_report.run_candidate_backtest_validation(
                pd.DataFrame(columns=["Rank", "Stock", "Signal", "Score"]),
                validate_top=5, strategy="ma_cross", period="1y", interval="1d",
                auto_adjust=False, force_refresh=False, initial_capital=100000,
                fee_rate=0.001425, tax_rate=0.003, position_size=1.0,
            )
            self.assertEqual(empty.columns.tolist(), daily_report.BACKTEST_HIGHLIGHT_COLUMNS)
            self.assertTrue(empty.empty)
            self.assertEqual(empty_limits, [])
            analyze.assert_not_called()
            run_backtest.assert_not_called()

        with self.assertRaisesRegex(ValueError, "Unsupported validation strategy"):
            daily_report.run_candidate_backtest_validation(
                self._candidates(1), validate_top=1, strategy="invalid", period="1y",
                interval="1d", auto_adjust=False, force_refresh=False, initial_capital=100000,
                fee_rate=0.001425, tax_rate=0.003, position_size=1.0,
            )

    def test_failure_continues_and_success_has_only_scalar_highlights(self) -> None:
        analysis = self._analysis()
        with patch.object(daily_report, "analyze_stock", side_effect=[ValueError("line one\nline two"), analysis]), patch.object(
            daily_report, "run_backtest_result", return_value=self._result()
        ), patch.dict(
            daily_report.STRATEGIES, {"score_strategy": lambda frame: frame}, clear=False
        ):
            highlights, limitations = daily_report.run_candidate_backtest_validation(
                self._candidates(2), validate_top=2, strategy="score", period="1y",
                interval="1d", auto_adjust=False, force_refresh=False, initial_capital=100000,
                fee_rate=0.001425, tax_rate=0.003, position_size=1.0,
            )

        self.assertEqual(highlights["Stock"].tolist(), ["2454", "2330"])
        self.assertEqual(highlights["Status"].tolist(), ["ERROR", "OK"])
        self.assertEqual(limitations, ["Backtest validation for 2454 failed: line one line two"])
        for value in highlights.iloc[1].tolist():
            self.assertNotIsInstance(value, (pd.DataFrame, pd.Series))
        self.assertEqual(highlights.columns.tolist(), daily_report.BACKTEST_HIGHLIGHT_COLUMNS)

class CandidateWalkForwardValidationTest(unittest.TestCase):
    def _backtests(self) -> pd.DataFrame:
        return pd.DataFrame([
            {"Rank": 2, "Stock": "2454", "Signal": "BUY", "Score": 6.2, "Status": "OK"},
            {"Rank": 1, "Stock": "2330", "Signal": "WATCH", "Score": 5.8, "Status": "ERROR"},
            {"Rank": 3, "Stock": "2317", "Signal": "BUY", "Score": 5.1, "Status": "OK"},
        ])

    def _detail(self, error: str = "") -> pd.DataFrame:
        return pd.DataFrame([
            {
                "Window": 1, "Test Total Return %": 10.123, "Test CAGR %": 8.5,
                "Test Sharpe Ratio": 1.234, "Test Max Drawdown %": -3.456, "Error": error,
            }
        ])

    def _run(self, detail: pd.DataFrame, **overrides):
        args = {
            "backtest_highlights": self._backtests(), "walk_forward_top": 2,
            "strategy": "score", "period": "2y", "interval": "1wk", "auto_adjust": True,
            "force_refresh": True, "train_days": 126, "test_days": 63, "step_days": None,
            "sort_by": "Train Sharpe Ratio", "initial_capital": 200000.0,
            "fee_rate": 0.001, "tax_rate": 0.002, "position_size": 0.5,
        }
        args.update(overrides)
        with patch.object(daily_report, "run_walk_forward", return_value=detail) as engine:
            result = daily_report.run_candidate_walk_forward_validation(**args)
        return result, engine

    def test_zero_is_noop_and_no_successful_backtests_skip(self) -> None:
        with patch.object(daily_report, "run_walk_forward") as engine:
            empty, limits = daily_report.run_candidate_walk_forward_validation(
                self._backtests(), walk_forward_top=0, strategy="score", period="1y",
                interval="1d", auto_adjust=False, force_refresh=False, train_days=126,
                test_days=63, step_days=None, sort_by="Train Sharpe Ratio",
                initial_capital=100000, fee_rate=0.001, tax_rate=0.003, position_size=1.0,
            )
        engine.assert_not_called()
        self.assertEqual(empty.columns.tolist(), daily_report.WALK_FORWARD_HIGHLIGHT_COLUMNS)
        self.assertEqual(limits, [])

        none = self._backtests().assign(Status="ERROR")
        with patch.object(daily_report, "run_walk_forward") as engine:
            empty, limits = daily_report.run_candidate_walk_forward_validation(
                none, walk_forward_top=1, strategy="score", period="1y", interval="1d",
                auto_adjust=False, force_refresh=False, train_days=126, test_days=63,
                step_days=None, sort_by="Train Sharpe Ratio", initial_capital=100000,
                fee_rate=0.001, tax_rate=0.003, position_size=1.0,
            )
        engine.assert_not_called()
        self.assertTrue(empty.empty)
        self.assertEqual(limits, ["Walk-forward validation skipped: no successful backtest candidates were available."])

    def test_order_limit_and_forwarding(self) -> None:
        details = self._detail()
        details = pd.concat([details, details.assign(Window=2, **{"Test Total Return %": 2.0})], ignore_index=True)
        (highlights, _), engine = self._run(details, walk_forward_top=1, step_days=None)
        self.assertEqual(highlights["Stock"].tolist(), ["2454"])
        self.assertEqual(highlights.loc[0, "Step Days"], 63)
        self.assertEqual(engine.call_args.kwargs, {
            "stock_id": "2454", "period": "2y", "strategy": "score", "train_days": 126,
            "test_days": 63, "step_days": 63, "sort_by": "Train Sharpe Ratio",
            "force_refresh": True, "position_size": 0.5, "initial_capital": 200000.0,
            "fee_rate": 0.001, "tax_rate": 0.002, "interval": "1wk", "auto_adjust": True,
        })

    def test_status_metrics_and_scalar_schema(self) -> None:
        detail = pd.DataFrame([
            {"Window": 1, "Test Total Return %": 10, "Test CAGR %": 8, "Test Sharpe Ratio": 1.2, "Test Max Drawdown %": -2, "Error": ""},
            {"Window": 2, "Test Total Return %": -2, "Test CAGR %": 1, "Test Sharpe Ratio": 0.5, "Test Max Drawdown %": -4, "Error": ""},
            {"Window": 3, "Test Total Return %": None, "Test CAGR %": None, "Test Sharpe Ratio": None, "Test Max Drawdown %": None, "Error": "line one\nline two"},
        ])
        (highlights, limits), _ = self._run(detail, walk_forward_top=1)
        row = highlights.iloc[0]
        self.assertEqual(row["Status"], "PARTIAL")
        self.assertEqual(row["Windows"], 3)
        self.assertEqual(row["Successful Windows"], 2)
        self.assertEqual(row["Error Windows"], 1)
        self.assertEqual(row["Positive Test Windows"], 1)
        self.assertEqual(row["Positive Test Windows %"], 50.0)
        self.assertEqual(row["Avg Test Total Return %"], 4.0)
        self.assertEqual(row["Best Test Sharpe Ratio"], 1.2)
        self.assertEqual(limits, ["Walk-forward validation for 2454 completed with 1 failed window(s): line one line two"])
        self.assertEqual(highlights.columns.tolist(), daily_report.WALK_FORWARD_HIGHLIGHT_COLUMNS)
        self.assertFalse(any(isinstance(value, (pd.DataFrame, pd.Series)) for value in row.tolist()))

        (ok, _), _ = self._run(self._detail())
        self.assertEqual(ok.loc[0, "Status"], "OK")
        (failed, limits), _ = self._run(self._detail("bad"), walk_forward_top=1)
        self.assertEqual(failed.loc[0, "Status"], "ERROR")
        self.assertEqual(failed.loc[0, "Avg Test Total Return %"], None)
        self.assertEqual(len(limits), 1)

    def test_candidate_failure_does_not_abort_later_candidate(self) -> None:
        details = self._detail()
        with patch.object(daily_report, "run_walk_forward", side_effect=[ValueError("first\nerror"), details]):
            highlights, limits = daily_report.run_candidate_walk_forward_validation(
                self._backtests(), walk_forward_top=2, strategy="score", period="1y", interval="1d",
                auto_adjust=False, force_refresh=False, train_days=10, test_days=5, step_days=5,
                sort_by="Train Sharpe Ratio", initial_capital=100000, fee_rate=0.001,
                tax_rate=0.003, position_size=1.0,
            )
        self.assertEqual(highlights["Status"].tolist(), ["ERROR", "OK"])
        self.assertEqual(limits, ["Walk-forward validation for 2454 failed: first error"])

    def test_unsupported_strategy_is_clear(self) -> None:
        with self.assertRaisesRegex(ValueError, "Unsupported walk-forward strategy"):
            daily_report.run_candidate_walk_forward_validation(
                self._backtests(), walk_forward_top=1, strategy="macd", period="1y",
                interval="1d", auto_adjust=False, force_refresh=False, train_days=10,
                test_days=5, step_days=5, sort_by="Train Sharpe Ratio", initial_capital=100000,
                fee_rate=0.001, tax_rate=0.003, position_size=1.0,
            )

    def test_markdown_renders_walk_forward_scalars(self) -> None:
        report = daily_report.build_daily_report_data(
            stock_universe=["2454"],
            walk_forward_highlights=pd.DataFrame([{
                "Rank": 1, "Stock": "2454", "Signal": "BUY", "Score": 6.2,
                "Strategy": "score", "Status": "OK", "Windows": 2,
                "Avg Test Total Return %": 4.0, "Best Test Sharpe Ratio": 1.2,
            }]),
        )
        markdown = daily_report.render_daily_report_markdown(report)
        self.assertIn("## Walk Forward Highlights", markdown)
        self.assertIn("2454", markdown)
        self.assertIn("OK", markdown)
        self.assertIn("4.0", markdown)
        self.assertNotIn("Equity Curve", markdown)
        self.assertNotIn("Trades", markdown)

if __name__ == "__main__":
    unittest.main()

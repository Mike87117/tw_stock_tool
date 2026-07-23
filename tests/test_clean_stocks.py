import contextlib
import io
from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from tw_stock_tool.cli import clean_stocks


def _download_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Open": [10.0, 11.0],
            "High": [12.0, 13.0],
            "Low": [9.0, 10.0],
            "Close": [11.0, 12.0],
            "Volume": [1000, 1100],
        },
        index=pd.date_range("2024-01-01", periods=2, freq="D"),
    )


def _fake_download(stock_id: str, **_: object) -> tuple[pd.DataFrame, str]:
    if stock_id == "9999":
        raise ValueError("bad stock")
    suffix = ".TWO" if stock_id in {"8069", "8069.TWO"} else ".TW"
    symbol = stock_id if stock_id.endswith((".TW", ".TWO")) else f"{stock_id}{suffix}"
    return _download_df(), symbol


class CleanStocksTest(unittest.TestCase):
    def _write_stock_file(self, text: str) -> tuple[tempfile.TemporaryDirectory, Path]:
        temp_dir = tempfile.TemporaryDirectory()
        path = Path(temp_dir.name) / "stocks.txt"
        path.write_text(text, encoding="utf-8")
        return temp_dir, path

    def test_read_stock_file_ignores_blank_and_comments(self) -> None:
        temp_dir, path = self._write_stock_file("\n# comment\n2330\n  \n2317\n")
        with temp_dir:
            entries, duplicates, total = clean_stocks.read_stock_file(path)

        self.assertEqual(total, 5)
        self.assertEqual([entry.normalized_stock for entry in entries], ["2330", "2317"])
        self.assertEqual(duplicates, [])

    def test_duplicate_stocks_checked_once_and_listed(self) -> None:
        temp_dir, path = self._write_stock_file("2330\n2317\n2330\n")
        with temp_dir:
            entries, duplicates, _ = clean_stocks.read_stock_file(path)
            duplicate_df = clean_stocks.duplicates_to_frame(duplicates)

        self.assertEqual([entry.normalized_stock for entry in entries], ["2330", "2317"])
        self.assertEqual(len(duplicate_df), 1)
        self.assertEqual(duplicate_df.loc[0, "Normalized Stock"], "2330")
        self.assertEqual(duplicate_df.loc[0, "First Row"], 1)

    def test_check_stock_entries_marks_ok_and_error(self) -> None:
        entries = [
            clean_stocks.StockEntry(1, "2330", "2330"),
            clean_stocks.StockEntry(2, "9999", "9999"),
        ]
        with patch.object(clean_stocks, "download_tw_stock", side_effect=_fake_download):
            result = clean_stocks.check_stock_entries(entries)

        self.assertEqual(result.loc[0, "Status"], "OK")
        self.assertEqual(result.loc[0, "Symbol"], "2330.TW")
        self.assertEqual(result.loc[1, "Status"], "ERROR")
        self.assertIn("bad stock", result.loc[1, "Error"])

    def test_summary_counts_are_correct(self) -> None:
        result = pd.DataFrame(
            [
                {"Status": "OK"},
                {"Status": "ERROR"},
                {"Status": "OK"},
            ]
        )
        duplicates = pd.DataFrame([{"Row": 4, "Stock": "2330"}])

        summary = clean_stocks.build_summary("stocks.txt", 5, result, duplicates, "clean.txt")

        self.assertEqual(summary.loc[0, "Total Input Lines"], 5)
        self.assertEqual(summary.loc[0, "Unique Stocks"], 3)
        self.assertEqual(summary.loc[0, "Valid Stocks"], 2)
        self.assertEqual(summary.loc[0, "Invalid Stocks"], 1)
        self.assertEqual(summary.loc[0, "Duplicate Rows"], 1)

    def test_export_excel_creates_expected_sheets(self) -> None:
        result = pd.DataFrame(
            [
                {
                    "Row": 1,
                    "Stock": "2330",
                    "Normalized Stock": "2330",
                    "Symbol": "2330.TW",
                    "Status": "OK",
                    "Error": "",
                    "Start Date": "2024-01-01",
                    "End Date": "2024-01-02",
                    "Rows": 2,
                    "Source Note": "Resolved symbol: 2330.TW",
                },
                {
                    "Row": 2,
                    "Stock": "9999",
                    "Normalized Stock": "9999",
                    "Symbol": "",
                    "Status": "ERROR",
                    "Error": "bad stock",
                    "Start Date": "",
                    "End Date": "",
                    "Rows": 0,
                    "Source Note": "",
                },
            ],
            columns=clean_stocks.RESULT_COLUMNS,
        )
        duplicates = pd.DataFrame(columns=clean_stocks.DUPLICATE_COLUMNS)
        summary = clean_stocks.build_summary("stocks.txt", 2, result, duplicates)

        with tempfile.TemporaryDirectory() as tmp_dir:
            output = Path(tmp_dir) / "clean.xlsx"
            path = clean_stocks.export_clean_report(summary, result, duplicates, output)
            workbook = load_workbook(path, read_only=True)
            sheetnames = workbook.sheetnames
            workbook.close()

        self.assertEqual(sheetnames, ["Summary", "Valid", "Invalid", "Duplicates", "All"])

    def test_write_clean_file_contains_only_valid_stocks(self) -> None:
        result = pd.DataFrame(
            [
                {"Normalized Stock": "2330", "Status": "OK"},
                {"Normalized Stock": "9999", "Status": "ERROR"},
                {"Normalized Stock": "8069", "Status": "OK"},
            ]
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            output = Path(tmp_dir) / "stocks_clean.txt"
            path = clean_stocks.write_clean_file(result, output)
            content = path.read_text(encoding="utf-8").splitlines()

        self.assertEqual(content, ["2330", "8069"])

    def test_explicit_market_suffix_is_preserved_in_clean_file(self) -> None:
        result = pd.DataFrame(
            [
                {"Normalized Stock": "2330.TW", "Status": "OK"},
                {"Normalized Stock": "8069.TWO", "Status": "OK"},
            ]
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            output = Path(tmp_dir) / "stocks_clean.txt"
            path = clean_stocks.write_clean_file(result, output)
            content = path.read_text(encoding="utf-8").splitlines()

        self.assertEqual(content, ["2330.TW", "8069.TWO"])


    def test_print_summary_uses_clear_english_output(self) -> None:
        result = pd.DataFrame(
            [
                {"Normalized Stock": "2330", "Status": "OK", "Error": ""},
                {"Normalized Stock": "9999", "Status": "ERROR", "Error": "bad stock"},
            ]
        )
        duplicates = pd.DataFrame([{"Row": 3, "Stock": "2330"}])
        summary = clean_stocks.build_summary("stocks.txt", 3, result, duplicates, "clean.txt")
        stream = io.StringIO()

        with contextlib.redirect_stdout(stream):
            clean_stocks.print_summary(
                summary,
                result,
                report_path=Path("output/clean_stocks_report.xlsx"),
                clean_path=Path("output/stocks_clean.txt"),
            )

        output = stream.getvalue()
        self.assertNotIn("???", output)
        self.assertIn("File:", output)
        self.assertIn("Total input lines:", output)
        self.assertIn("Unique stocks:", output)
        self.assertIn("Valid stocks:", output)
        self.assertIn("Invalid stocks:", output)
        self.assertIn("Duplicate rows:", output)

    def test_run_clean_stocks_uses_mocked_download_and_outputs(self) -> None:
        temp_dir, stock_file = self._write_stock_file("# comment\n2330\n8069\n9999\n2330\n")
        with temp_dir:
            report_path = Path(temp_dir.name) / "report.xlsx"
            clean_path = Path(temp_dir.name) / "clean.txt"
            with patch.object(clean_stocks, "download_tw_stock", side_effect=_fake_download) as download:
                summary, result, duplicates, report, clean_file = clean_stocks.run_clean_stocks(
                    stock_file,
                    output=report_path,
                    clean_file=clean_path,
                )

            self.assertEqual(download.call_count, 3)
            self.assertEqual(summary.loc[0, "Total Input Lines"], 5)
            self.assertEqual(summary.loc[0, "Valid Stocks"], 2)
            self.assertEqual(summary.loc[0, "Invalid Stocks"], 1)
            self.assertEqual(summary.loc[0, "Duplicate Rows"], 1)
            self.assertEqual(len(duplicates), 1)
            self.assertTrue(report.exists())
            self.assertTrue(clean_file.exists())
            self.assertEqual(clean_file.read_text(encoding="utf-8").splitlines(), ["2330", "8069"])
            self.assertEqual(result[result["Status"] == "OK"]["Stock"].tolist(), ["2330", "8069"])


if __name__ == "__main__":
    unittest.main()

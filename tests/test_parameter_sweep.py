import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

import parameter_sweep
from analysis import StockAnalysis


def _sample_signal_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Close": [10.0, 11.0, 12.0, 11.0, 13.0],
            "Signal": ["HOLD", "BUY", "HOLD", "SELL", "HOLD"],
            "Score": [0.0, 5.0, 2.0, -3.0, 1.0],
            "MA5": [9.0, 10.0, 11.0, 10.0, 12.0],
            "MA20": [10.0, 10.0, 10.0, 10.0, 10.0],
            "MACD": [0.0, 1.0, 1.2, 0.5, 0.8],
            "MACD_Signal": [0.5, 0.8, 1.0, 0.8, 0.7],
            "RSI": [50.0, 25.0, 35.0, 75.0, 55.0],
        },
        index=pd.date_range("2024-01-01", periods=5, freq="D"),
    )


def _sample_sweep_result() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Rank": 1,
                "Strategy": "ma_cross",
                "Parameters": "short_window=5, long_window=20",
                "Total Return %": 10.0,
                "Buy and Hold Return %": 8.0,
                "CAGR %": 4.0,
                "Trade Count": 2,
                "Win Rate %": 50.0,
                "Max Drawdown %": -3.0,
                "Profit Factor": 1.5,
                "Sharpe Ratio": 0.8,
                "Sortino Ratio": 0.9,
                "Error": "",
            },
            {
                "Rank": 2,
                "Strategy": "rsi",
                "Parameters": "buy_below=30, sell_above=70",
                "Total Return %": 7.0,
                "Buy and Hold Return %": 8.0,
                "CAGR %": 3.0,
                "Trade Count": 1,
                "Win Rate %": 100.0,
                "Max Drawdown %": -2.0,
                "Profit Factor": 2.0,
                "Sharpe Ratio": 0.7,
                "Sortino Ratio": 0.8,
                "Error": "",
            },
            {
                "Rank": 3,
                "Strategy": "score",
                "Parameters": "buy_score=4, sell_score=-2",
                "Total Return %": 5.0,
                "Buy and Hold Return %": 8.0,
                "CAGR %": 2.0,
                "Trade Count": 1,
                "Win Rate %": 0.0,
                "Max Drawdown %": -4.0,
                "Profit Factor": 0.5,
                "Sharpe Ratio": 0.4,
                "Sortino Ratio": 0.5,
                "Error": "",
            },
            {
                "Rank": None,
                "Strategy": "score",
                "Parameters": "buy_score=6, sell_score=-4",
                "Total Return %": None,
                "Buy and Hold Return %": None,
                "CAGR %": None,
                "Trade Count": None,
                "Win Rate %": None,
                "Max Drawdown %": None,
                "Profit Factor": None,
                "Sharpe Ratio": None,
                "Sortino Ratio": None,
                "Error": "bad params",
            },
        ],
        columns=parameter_sweep.SWEEP_COLUMNS,
    )


def _fake_analysis() -> StockAnalysis:
    signal_df = _sample_signal_df()
    return StockAnalysis(
        stock_id="2330",
        symbol="2330.TW",
        raw_df=pd.DataFrame(),
        indicator_df=pd.DataFrame(),
        signal_df=signal_df,
        latest=signal_df.iloc[-1],
        summary={},
    )


def _fake_backtest(df: pd.DataFrame, **_: object) -> dict[str, object]:
    signal_score = {"BUY": 2.0, "SELL": -1.0, "HOLD": 0.0}
    total_return = sum(signal_score.get(signal, 0.0) for signal in df["Signal"])
    return {
        "Total Return %": total_return,
        "Buy and Hold Return %": 1.0,
        "CAGR %": total_return / 10,
        "Trade Count": int((df["Signal"] == "BUY").sum()),
        "Win Rate %": 50.0,
        "Max Drawdown %": -5.0,
        "Profit Factor": 1.2,
        "Sharpe Ratio": 0.5,
        "Sortino Ratio": 0.4,
    }


class ParameterSweepTest(unittest.TestCase):
    def test_ma_parameter_grid_keeps_short_less_than_long(self) -> None:
        grid = parameter_sweep.ma_cross_parameter_grid()

        self.assertEqual(len(grid), 8)
        self.assertTrue(all(item["short_window"] < item["long_window"] for item in grid))

    def test_rsi_parameter_grid_keeps_buy_less_than_sell(self) -> None:
        grid = parameter_sweep.rsi_parameter_grid()

        self.assertEqual(len(grid), 9)
        self.assertTrue(all(item["buy_below"] < item["sell_above"] for item in grid))

    def test_score_parameter_grid_keeps_buy_greater_than_sell(self) -> None:
        grid = parameter_sweep.score_parameter_grid()

        self.assertEqual(len(grid), 9)
        self.assertTrue(all(item["buy_score"] > item["sell_score"] for item in grid))

    def test_parameter_sweep_returns_dataframe(self) -> None:
        with patch.object(parameter_sweep, "analyze_stock", return_value=_fake_analysis()):
            with patch.object(parameter_sweep, "run_backtest", side_effect=_fake_backtest):
                result = parameter_sweep.run_parameter_sweep(
                    "2330",
                    strategy="score",
                    top=5,
                )

        self.assertIsInstance(result, pd.DataFrame)
        self.assertEqual(len(result), 5)
        self.assertIn("Rank", result.columns)
        self.assertIn("Total Return %", result.columns)

    def test_single_strategy_failure_does_not_stop_sweep(self) -> None:
        original_score_strategy = parameter_sweep.score_strategy

        def flaky_score_strategy(df: pd.DataFrame, **params: int) -> pd.DataFrame:
            if params["buy_score"] == 4 and params["sell_score"] == -2:
                raise ValueError("bad params")
            return original_score_strategy(df, **params)

        with patch.object(parameter_sweep, "analyze_stock", return_value=_fake_analysis()):
            with patch.object(parameter_sweep, "run_backtest", side_effect=_fake_backtest):
                with patch.object(parameter_sweep, "score_strategy", side_effect=flaky_score_strategy):
                    result = parameter_sweep.run_parameter_sweep("2330", strategy="score", top=20)

        self.assertEqual(len(result), 9)
        self.assertTrue(result["Error"].astype(str).str.contains("bad params").any())
        self.assertGreater((result["Error"] == "").sum(), 0)

    def test_sort_by_missing_column_raises(self) -> None:
        with patch.object(parameter_sweep, "analyze_stock", return_value=_fake_analysis()):
            with patch.object(parameter_sweep, "run_backtest", side_effect=_fake_backtest):
                with self.assertRaises(ValueError):
                    parameter_sweep.run_parameter_sweep("2330", sort_by="Missing")

    def test_sort_by_supports_all_sortable_columns(self) -> None:
        with patch.object(parameter_sweep, "analyze_stock", return_value=_fake_analysis()):
            with patch.object(parameter_sweep, "run_backtest", side_effect=_fake_backtest):
                for column in parameter_sweep.SORTABLE_COLUMNS:
                    with self.subTest(column=column):
                        result = parameter_sweep.run_parameter_sweep(
                            "2330",
                            strategy="score",
                            sort_by=column,
                            top=3,
                        )
                        self.assertFalse(result.empty)

    def test_sort_by_rejects_text_columns(self) -> None:
        with patch.object(parameter_sweep, "analyze_stock", return_value=_fake_analysis()):
            with patch.object(parameter_sweep, "run_backtest", side_effect=_fake_backtest):
                for column in ["Parameters", "Strategy", "Error"]:
                    with self.subTest(column=column):
                        with self.assertRaisesRegex(ValueError, "Supported columns"):
                            parameter_sweep.run_parameter_sweep(
                                "2330",
                                strategy="score",
                                sort_by=column,
                            )

    def test_top_zero_returns_all_success_and_error_rows(self) -> None:
        result = self._run_sweep_with_one_error(top=0)

        self.assertEqual(len(result), 9)
        self.assertEqual((result["Error"] == "").sum(), 8)
        self.assertEqual((result["Error"] != "").sum(), 1)

    def test_top_negative_returns_all_success_and_error_rows(self) -> None:
        result = self._run_sweep_with_one_error(top=-1)

        self.assertEqual(len(result), 9)
        self.assertEqual((result["Error"] == "").sum(), 8)
        self.assertEqual((result["Error"] != "").sum(), 1)


    def test_export_parameter_sweep_csv_creates_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "sweep.csv"
            result = parameter_sweep.export_parameter_sweep(
                _sample_sweep_result(),
                "2330",
                str(output_path),
            )

            self.assertEqual(result, output_path)
            self.assertTrue(output_path.exists())
            self.assertIn("Strategy", output_path.read_text(encoding="utf-8-sig"))

    def test_export_parameter_sweep_excel_creates_xlsx(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "sweep.xlsx"
            result = parameter_sweep.export_parameter_sweep_excel(
                _sample_sweep_result(),
                "2330",
                str(output_path),
            )

            self.assertEqual(result, output_path)
            self.assertTrue(output_path.exists())

    def test_excel_contains_all_and_errors_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "sweep.xlsx"
            parameter_sweep.export_parameter_sweep_excel(
                _sample_sweep_result(),
                "2330",
                str(output_path),
            )
            workbook = load_workbook(output_path, read_only=True)

            self.assertIn("All", workbook.sheetnames)
            self.assertIn("Errors", workbook.sheetnames)
            workbook.close()

    def test_excel_splits_by_strategy(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "sweep.xlsx"
            parameter_sweep.export_parameter_sweep_excel(
                _sample_sweep_result(),
                "2330",
                str(output_path),
            )
            workbook = load_workbook(output_path, read_only=True)

            self.assertIn("MA_Cross", workbook.sheetnames)
            self.assertIn("RSI", workbook.sheetnames)
            self.assertIn("Score", workbook.sheetnames)
            workbook.close()

    def test_output_excel_default_path(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            with patch.object(parameter_sweep, "OUTPUT_DIR", Path(tmp_dir)):
                result = parameter_sweep.export_parameter_sweep_excel(
                    _sample_sweep_result(),
                    "2330",
                    "",
                )

            self.assertEqual(result, Path(tmp_dir) / "2330_parameter_sweep.xlsx")
            self.assertTrue(result.exists())

    def test_output_excel_custom_path(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "custom_parameter_sweep.xlsx"
            result = parameter_sweep.export_parameter_sweep_excel(
                _sample_sweep_result(),
                "2330",
                str(output_path),
            )

            self.assertEqual(result, output_path)
            self.assertTrue(output_path.exists())


    def test_parameter_sweep_uses_next_day_backtest_execution(self) -> None:
        signal_df = pd.DataFrame(
            {
                "Close": [100.0, 110.0, 120.0, 130.0],
                "Signal": ["HOLD", "HOLD", "HOLD", "HOLD"],
                "Score": [5.0, 0.0, -3.0, 0.0],
            },
            index=pd.date_range("2024-01-01", periods=4, freq="D"),
        )
        analysis = StockAnalysis(
            stock_id="2330",
            symbol="2330.TW",
            raw_df=pd.DataFrame(),
            indicator_df=pd.DataFrame(),
            signal_df=signal_df,
            latest=signal_df.iloc[-1],
            summary={},
        )

        with patch.object(parameter_sweep, "analyze_stock", return_value=analysis):
            with patch.object(parameter_sweep, "INITIAL_CAPITAL", 10000):
                with patch.object(parameter_sweep, "FEE_RATE", 0):
                    with patch.object(parameter_sweep, "TAX_RATE", 0):
                        result = parameter_sweep.run_parameter_sweep(
                            "2330",
                            strategy="score",
                            top=20,
                        )

        target = result[result["Parameters"] == "buy_score=4, sell_score=-2"].iloc[0]
        # Score creates BUY on day 1 and SELL on day 3. run_backtest must
        # execute those orders at day 2 and day 4 closes, yielding 18%.
        self.assertEqual(target["Total Return %"], 18.0)

    def test_invalid_position_size_raises(self) -> None:
        with self.assertRaises(ValueError):
            parameter_sweep.run_parameter_sweep("2330", position_size=0)

    def _run_sweep_with_one_error(self, top: int) -> pd.DataFrame:
        original_score_strategy = parameter_sweep.score_strategy

        def flaky_score_strategy(df: pd.DataFrame, **params: int) -> pd.DataFrame:
            if params["buy_score"] == 4 and params["sell_score"] == -2:
                raise ValueError("bad params")
            return original_score_strategy(df, **params)

        with patch.object(parameter_sweep, "analyze_stock", return_value=_fake_analysis()):
            with patch.object(parameter_sweep, "run_backtest", side_effect=_fake_backtest):
                with patch.object(parameter_sweep, "score_strategy", side_effect=flaky_score_strategy):
                    return parameter_sweep.run_parameter_sweep(
                        "2330",
                        strategy="score",
                        top=top,
                    )


if __name__ == "__main__":
    unittest.main()

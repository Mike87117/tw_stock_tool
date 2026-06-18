from pathlib import Path
import tempfile
import unittest

import pandas as pd
from openpyxl import load_workbook

from report import ReportError, export_stock_ranking


class ReportTest(unittest.TestCase):
    def test_export_stock_ranking_creates_all_formats(self) -> None:
        with self.subTest("all formats"):
            with tempfile.TemporaryDirectory() as tmp_dir:
                df = pd.DataFrame(
                    [
                        {
                            "Rank": 1,
                            "Stock": "2330",
                            "Signal": "BUY",
                            "Score": 5.5,
                            "Status": "OK",
                        }
                    ]
                )
                paths = export_stock_ranking(df, Path(tmp_dir))

                self.assertTrue(paths["excel"].exists())
                self.assertTrue(paths["csv"].exists())
                self.assertTrue(paths["html"].exists())

    def test_export_stock_ranking_rejects_missing_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            with self.assertRaisesRegex(ReportError, "排行榜資料不可缺少欄位"):
                export_stock_ranking(pd.DataFrame(), Path(tmp_dir))

    def test_export_stock_ranking_allows_empty_result_with_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            df = pd.DataFrame(columns=["Rank", "Stock", "Signal", "Score", "Status"])
            paths = export_stock_ranking(df, Path(tmp_dir))
            self.assertTrue(paths["excel"].exists())

    def test_export_stock_ranking_sheet_by_signal(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            df = pd.DataFrame(
                [
                    {"Rank": 1, "Stock": "2330", "Signal": "BUY", "Score": 4, "Status": "OK"},
                    {"Rank": 2, "Stock": "2317", "Signal": "WATCH", "Score": 2, "Status": "OK"},
                ]
            )

            paths = export_stock_ranking(df, Path(tmp_dir), sheet_by_signal=True)
            wb = load_workbook(paths["excel"], read_only=True)
            sheetnames = wb.sheetnames
            wb.close()

        self.assertIn("Ranking", sheetnames)
        self.assertIn("BUY", sheetnames)
        self.assertIn("WATCH", sheetnames)


if __name__ == "__main__":
    unittest.main()

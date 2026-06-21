import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

import ai_stock_scanner


def _summary(stock: str, f1: float, accuracy: float, errors: int = 0) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Stock": stock,
                "Period": "5y",
                "Horizon": 5,
                "Train Size": 252,
                "Test Size": 63,
                "Step Size": 63,
                "Windows": 3,
                "Avg Accuracy": accuracy,
                "Avg Precision": 0.5,
                "Avg Recall": 0.6,
                "Avg F1": f1,
                "Avg Test Positive Rate %": 52.0,
                "Avg Predicted Positive Rate %": 48.0,
                "Error Windows": errors,
            }
        ]
    )


def _frames(stock: str, f1: float, accuracy: float, errors: int = 0) -> dict[str, pd.DataFrame]:
    return {
        "Summary": _summary(stock, f1, accuracy, errors),
        "Detail": pd.DataFrame(),
        "Errors": pd.DataFrame(),
    }


class AIStockScannerTest(unittest.TestCase):
    def test_collect_stock_ids_from_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "stocks.txt"
            path.write_text("2330\n# comment\n\n2317\n2330\n", encoding="utf-8")
            result = ai_stock_scanner.collect_stock_ids(file_path=path)

        self.assertEqual(result, ["2330", "2317"])

    def test_collect_stock_ids_from_stocks(self) -> None:
        result = ai_stock_scanner.collect_stock_ids(stocks=["2330", "", "2317", "2330"])

        self.assertEqual(result, ["2330", "2317"])

    def test_scan_uses_mocked_report_without_network(self) -> None:
        def fake_report(stock_id: str, **_: object) -> dict[str, pd.DataFrame]:
            return _frames(stock_id, f1=0.5, accuracy=0.6)

        with patch.object(ai_stock_scanner, "run_ai_prediction_report", side_effect=fake_report) as mocked:
            result = ai_stock_scanner.scan_ai_stocks(
                ["2330", "2317"],
                period="5y",
                horizon=5,
                workers=1,
            )

        self.assertEqual(mocked.call_count, 2)
        self.assertEqual(set(result["Status"]), {"OK"})
        self.assertEqual(len(result), 2)

    def test_single_stock_failure_does_not_stop_batch(self) -> None:
        def fake_report(stock_id: str, **_: object) -> dict[str, pd.DataFrame]:
            if stock_id == "9999":
                raise ValueError("bad stock")
            return _frames(stock_id, f1=0.5, accuracy=0.6)

        with patch.object(ai_stock_scanner, "run_ai_prediction_report", side_effect=fake_report):
            result = ai_stock_scanner.scan_ai_stocks(["2330", "9999"], workers=1)

        self.assertEqual(len(result), 2)
        self.assertEqual(result.iloc[0]["Status"], "OK")
        self.assertEqual(result.iloc[1]["Status"], "ERROR")
        self.assertIn("bad stock", result.iloc[1]["Error"])

    def test_ranking_sorts_by_f1_accuracy_and_error_windows(self) -> None:
        rows = [
            ai_stock_scanner._summary_to_row("low", _summary("low", f1=0.4, accuracy=0.9, errors=0)),
            ai_stock_scanner._summary_to_row("best", _summary("best", f1=0.8, accuracy=0.7, errors=1)),
            ai_stock_scanner._summary_to_row("tie", _summary("tie", f1=0.8, accuracy=0.8, errors=2)),
            ai_stock_scanner._summary_to_row("clean", _summary("clean", f1=0.8, accuracy=0.8, errors=0)),
        ]

        result = ai_stock_scanner.rank_ai_stock_results(rows)

        self.assertEqual(result["Stock"].tolist(), ["clean", "tie", "best", "low"])
        self.assertEqual(result["Rank"].tolist(), [1, 2, 3, 4])

    def test_excel_output_can_be_created(self) -> None:
        ranking = ai_stock_scanner.rank_ai_stock_results(
            [ai_stock_scanner._summary_to_row("2330", _summary("2330", 0.7, 0.8))]
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "ranking.xlsx"
            result = ai_stock_scanner.export_ai_stock_ranking(ranking, str(output_path))
            workbook = load_workbook(result, read_only=True)

            self.assertEqual(result, output_path)
            self.assertIn("Ranking", workbook.sheetnames)
            workbook.close()

    def test_collect_stock_ids_auto_stock_list_calls_updater_and_has_priority(self) -> None:
        updater_df = pd.DataFrame([{"Stock": "2330"}, {"Stock": "2317"}])
        with patch.object(
            ai_stock_scanner.stock_list_updater_module,
            "update_stock_list",
            return_value=(updater_df, "stocks.txt"),
        ) as updater_mock:
            result = ai_stock_scanner.collect_stock_ids(
                stocks=["9999"],
                file_path="ignored.txt",
                auto_stock_list=True,
                stock_market="all",
                stock_list_output="stocks.txt",
                allow_partial_stock_list=True,
            )

        updater_mock.assert_called_once_with(
            market="all",
            output="stocks.txt",
            allow_partial=True,
        )
        self.assertEqual(result, ["2330", "2317"])

    def test_collect_stock_ids_applies_stock_limit(self) -> None:
        result = ai_stock_scanner.collect_stock_ids(
            stocks=["2330", "2317", "2454"],
            stock_limit=2,
        )

        self.assertEqual(result, ["2330", "2317"])

    def test_collect_stock_ids_applies_stock_sample(self) -> None:
        first = ai_stock_scanner.collect_stock_ids(
            stocks=["2330", "2317", "2454", "2308"],
            stock_sample=2,
            random_state=7,
        )
        second = ai_stock_scanner.collect_stock_ids(
            stocks=["2330", "2317", "2454", "2308"],
            stock_sample=2,
            random_state=7,
        )

        self.assertEqual(first, second)
        self.assertEqual(len(first), 2)

    def test_parse_args_supports_auto_stock_list(self) -> None:
        args = ai_stock_scanner._parse_args([
            "--auto-stock-list",
            "--stock-market",
            "twse",
            "--stock-list-output",
            "stocks.txt",
            "--allow-partial-stock-list",
        ])

        self.assertTrue(args.auto_stock_list)
        self.assertEqual(args.stock_market, "twse")
        self.assertEqual(args.stock_list_output, "stocks.txt")
        self.assertTrue(args.allow_partial_stock_list)

    def test_ai_stock_scanner_updater_failure_does_not_scan(self) -> None:
        args = ai_stock_scanner._parse_args(["--auto-stock-list"])
        with patch.object(ai_stock_scanner, "_parse_args", return_value=args):
            with patch.object(
                ai_stock_scanner.stock_list_updater_module,
                "update_stock_list",
                side_effect=RuntimeError("updater down"),
            ):
                with patch.object(ai_stock_scanner, "scan_ai_stocks") as scan_mock:
                    with patch("builtins.print"):
                        ai_stock_scanner.main()

        scan_mock.assert_not_called()

    def test_parse_args(self) -> None:
        args = ai_stock_scanner._parse_args(
            [
                "--stocks",
                "2330",
                "2317",
                "--period",
                "5y",
                "--horizon",
                "5",
                "--train-size",
                "252",
                "--test-size",
                "63",
                "--workers",
                "2",
                "--output",
            ]
        )

        self.assertEqual(args.stocks, ["2330", "2317"])
        self.assertEqual(args.period, "5y")
        self.assertEqual(args.horizon, 5)
        self.assertEqual(args.train_size, 252)
        self.assertEqual(args.test_size, 63)
        self.assertEqual(args.workers, 2)
        self.assertEqual(args.output, "")


if __name__ == "__main__":
    unittest.main()

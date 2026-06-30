import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook
from typing import Any
import unittest.mock as mock

import walk_forward
from analysis import StockAnalysis


class WalkForwardTest(unittest.TestCase):
    def _sample_df(self, rows: int = 80) -> pd.DataFrame:
        index = pd.date_range("2024-01-01", periods=rows, freq="B")
        close = pd.Series([100 + (i % 17) + i * 0.2 for i in range(rows)], index=index)
        df = pd.DataFrame(
            {
                "Close": close,
                "Signal": ["BUY" if i % 11 == 0 else "SELL" if i % 17 == 0 else "HOLD" for i in range(rows)],
                "Score": [6 if i % 9 == 0 else -3 if i % 13 == 0 else 1 for i in range(rows)],
                "RSI": [20 if i % 10 == 0 else 80 if i % 14 == 0 else 50 for i in range(rows)],
                "MACD": [i * 0.01 for i in range(rows)],
                "MACD_Signal": [i * 0.008 for i in range(rows)],
            },
            index=index,
        )
        return df

    def _analysis(self, rows: int = 80) -> StockAnalysis:
        df = self._sample_df(rows)
        return StockAnalysis(
            stock_id="2330",
            symbol="2330.TW",
            raw_df=df,
            indicator_df=df,
            signal_df=df,
            latest=df.iloc[-1],
            summary={},
        )

    def _fake_backtest(self, df: pd.DataFrame, **_: object) -> dict[str, object]:
        buys = int((df["Signal"] == "BUY").sum())
        sells = int((df["Signal"] == "SELL").sum())
        total_return = float(buys - sells)
        return {
            "Total Return %": total_return,
            "Buy and Hold Return %": 1.0,
            "CAGR %": total_return / 2,
            "Trade Count": buys,
            "Win Rate %": 50.0 if buys else 0.0,
            "Max Drawdown %": -float(sells),
            "Profit Factor": 1.5 + buys,
            "Sharpe Ratio": total_return + 0.5,
            "Sortino Ratio": total_return + 0.25,
            "Trades": pd.DataFrame(),
            "Equity Curve": pd.Series(dtype=float),
        }

    def test_window_splitting(self) -> None:
        df = self._sample_df(10)
        windows = walk_forward.split_windows(df, train_days=4, test_days=2, step_days=2)

        self.assertEqual(len(windows), 3)
        self.assertEqual(windows[0][0], 1)
        self.assertEqual(windows[0][1].index[0], df.index[0])
        self.assertEqual(windows[0][2].index[0], df.index[4])


    def test_train_test_windows_do_not_overlap(self) -> None:
        df = self._sample_df(12)
        windows = walk_forward.split_windows(df, train_days=4, test_days=3, step_days=2)

        for _, train, test in windows:
            self.assertTrue(set(train.index).isdisjoint(set(test.index)))
            self.assertLess(train.index[-1], test.index[0])

    def test_window_splitting_requires_enough_data(self) -> None:
        df = self._sample_df(5)

        with self.assertRaises(ValueError):
            walk_forward.split_windows(df, train_days=4, test_days=2, step_days=1)

    def test_ma_cross_walk_forward(self) -> None:
        with patch("walk_forward.analyze_stock", return_value=self._analysis()), patch(
            "walk_forward.run_backtest",
            side_effect=self._fake_backtest,
        ):
            result = walk_forward.run_walk_forward(
                "2330",
                strategy="ma_cross",
                train_days=20,
                test_days=10,
                step_days=10,
            )

        self.assertFalse(result.empty)
        self.assertEqual(set(result["Strategy"]), {"ma_cross"})

    def test_rsi_walk_forward(self) -> None:
        with patch("walk_forward.analyze_stock", return_value=self._analysis()), patch(
            "walk_forward.run_backtest",
            side_effect=self._fake_backtest,
        ):
            result = walk_forward.run_walk_forward(
                "2330",
                strategy="rsi",
                train_days=20,
                test_days=10,
                step_days=10,
            )

        self.assertFalse(result.empty)
        self.assertEqual(set(result["Strategy"]), {"rsi"})

    def test_score_walk_forward(self) -> None:
        with patch("walk_forward.analyze_stock", return_value=self._analysis()), patch(
            "walk_forward.run_backtest",
            side_effect=self._fake_backtest,
        ):
            result = walk_forward.run_walk_forward(
                "2330",
                strategy="score",
                train_days=20,
                test_days=10,
                step_days=10,
            )

        self.assertFalse(result.empty)
        self.assertEqual(set(result["Strategy"]), {"score"})

    def test_window_failure_does_not_stop_all_results(self) -> None:
        with patch("walk_forward.analyze_stock", return_value=self._analysis()), patch(
            "walk_forward.run_backtest",
            side_effect=self._fake_backtest,
        ), patch("walk_forward.ma_cross_strategy", side_effect=ValueError("strategy failed")):
            result = walk_forward.run_walk_forward(
                "2330",
                strategy="all",
                train_days=20,
                test_days=10,
                step_days=10,
            )

        self.assertGreater((result["Error"].astype(str) != "").sum(), 0)
        self.assertGreater((result["Error"].astype(str) == "").sum(), 0)


    def test_parameter_selection_uses_train_metrics_only(self) -> None:
        df = self._sample_df(8)
        train = df.iloc[:4].copy()
        test = df.iloc[4:].copy()
        params_a = {"buy_score": 4, "sell_score": -2}
        params_b = {"buy_score": 6, "sell_score": -4}
        test_calls: list[dict[str, int]] = []

        def fake_backtest(
            data: pd.DataFrame,
            strategy: str,
            params: dict[str, int],
            *_: object,
        ) -> dict[str, object]:
            is_train = data.index[0] == train.index[0]
            if not is_train:
                test_calls.append(params.copy())
            sharpe = 10.0 if is_train and params == params_a else 1.0
            if not is_train and params == params_b:
                sharpe = 999.0
            return {
                "Total Return %": sharpe,
                "CAGR %": sharpe,
                "Trade Count": 1,
                "Win Rate %": 50.0,
                "Max Drawdown %": -1.0,
                "Profit Factor": sharpe,
                "Sharpe Ratio": sharpe,
                "Sortino Ratio": sharpe,
            }

        with patch.object(walk_forward, "_parameter_grid", return_value=[params_a, params_b]):
            with patch.object(walk_forward, "_run_strategy_backtest", side_effect=fake_backtest):
                result = walk_forward._evaluate_window_strategy(
                    window_number=1,
                    train=train,
                    test=test,
                    strategy="score",
                    sort_by="Train Sharpe Ratio",
                    stop_loss_pct=None,
                    take_profit_pct=None,
                    max_hold_days=None,
                    position_size=1.0,
                    initial_capital=100000.0,
                    fee_rate=0.001425,
                    tax_rate=0.003,
                )

        self.assertEqual(result["Parameters"], "buy_score=4, sell_score=-2")
        self.assertEqual(test_calls, [params_a])

    @mock.patch("src.tw_stock_tool.backtesting.walk_forward.analyze_stock")
    @mock.patch("src.tw_stock_tool.backtesting.walk_forward._run_strategy_backtest")
    def test_walk_forward_custom_ranges(
        self, mock_run_backtest: mock.MagicMock, mock_analyze: mock.MagicMock
    ) -> None:
        mock_analyze.return_value.signal_df = self._sample_df(rows=100)

        # Provide deterministic fake results for all params
        def side_effect(df: pd.DataFrame, strategy: str, params: dict[str, int], *args: Any, **kwargs: Any) -> dict[str, Any]:
            return {
                "Total Return %": 10.0,
                "CAGR %": 10.0,
                "Trade Count": 2,
                "Win Rate %": 50.0,
                "Max Drawdown %": -5.0,
                "Profit Factor": 1.5,
                "Sharpe Ratio": 1.0,
                "Sortino Ratio": 1.0,
            }
        mock_run_backtest.side_effect = side_effect

        df = walk_forward.run_walk_forward(
            stock_id="2330",
            period="1y",
            strategy="ma_cross",
            train_days=50,
            test_days=20,
            ma_short_windows=(2, 3),
            ma_long_windows=(3, 4)
        )

        # Expected parameters in grid:
        # short=2, long=3
        # short=2, long=4
        # short=3, long=4
        # Test just the first window's best param
        self.assertFalse(df.empty)
        # Verify custom ranges didn't fail
        for param in df["Parameters"]:
            self.assertTrue("short_window=2" in param or "short_window=3" in param)

    @mock.patch("src.tw_stock_tool.backtesting.walk_forward.analyze_stock")
    @mock.patch.object(walk_forward, "run_backtest")
    @mock.patch("src.tw_stock_tool.backtesting.walk_forward._build_strategy_df")
    def test_run_walk_forward_engine_params_passthrough(
        self, mock_build_strategy: mock.MagicMock, mock_run_backtest: mock.MagicMock, mock_analyze: mock.MagicMock
    ) -> None:
        mock_analyze.return_value.signal_df = self._sample_df(rows=200)
        mock_build_strategy.return_value = pd.DataFrame({"Close": [1], "Signal": [1]})

        mock_run_backtest.return_value = {
            "Total Return %": 10.0,
            "CAGR %": 10.0,
            "Trade Count": 2,
            "Win Rate %": 50.0,
            "Max Drawdown %": -5.0,
            "Profit Factor": 1.5,
            "Sharpe Ratio": 1.0,
            "Sortino Ratio": 1.0,
        }

        df = walk_forward.run_walk_forward(
            stock_id="2330",
            strategy="ma_cross",
            train_days=50,
            test_days=20,
            initial_capital=99999.0,
            fee_rate=0.01,
            tax_rate=0.02,
            position_size=0.5,
            stop_loss_pct=0.05,
            take_profit_pct=0.1,
            max_hold_days=10,
        )

        self.assertTrue(mock_run_backtest.called)
        # Check the last call to ensure parameters were passed
        _, kwargs = mock_run_backtest.call_args
        self.assertEqual(kwargs["initial_capital"], 99999.0)
        self.assertEqual(kwargs["fee_rate"], 0.01)
        self.assertEqual(kwargs["tax_rate"], 0.02)
        self.assertEqual(kwargs["position_size"], 0.5)
        self.assertEqual(kwargs["stop_loss_pct"], 0.05)
        self.assertEqual(kwargs["take_profit_pct"], 0.1)
        self.assertEqual(kwargs["max_hold_days"], 10)

    def test_export_excel_contains_required_sheets(self) -> None:
        detail = pd.DataFrame(
            [
                {
                    "Window": 1,
                    "Train Start": pd.Timestamp("2024-01-01"),
                    "Train End": pd.Timestamp("2024-01-31"),
                    "Test Start": pd.Timestamp("2024-02-01"),
                    "Test End": pd.Timestamp("2024-02-15"),
                    "Strategy": "score",
                    "Parameters": "buy_score=4, sell_score=-2",
                    "Train Total Return %": 5.0,
                    "Test Total Return %": 1.0,
                    "Train CAGR %": 8.0,
                    "Test CAGR %": 2.0,
                    "Train Trade Count": 2,
                    "Test Trade Count": 1,
                    "Train Win Rate %": 50.0,
                    "Test Win Rate %": 100.0,
                    "Train Max Drawdown %": -3.0,
                    "Test Max Drawdown %": -1.0,
                    "Train Profit Factor": 2.0,
                    "Test Profit Factor": 1.2,
                    "Train Sharpe Ratio": 1.1,
                    "Test Sharpe Ratio": 0.8,
                    "Train Sortino Ratio": 1.3,
                    "Test Sortino Ratio": 0.9,
                    "Error": "",
                },
                {
                    "Window": 2,
                    "Train Start": pd.Timestamp("2024-02-01"),
                    "Train End": pd.Timestamp("2024-02-28"),
                    "Test Start": pd.Timestamp("2024-03-01"),
                    "Test End": pd.Timestamp("2024-03-15"),
                    "Strategy": "score",
                    "Parameters": "",
                    "Error": "failed",
                },
            ],
            columns=walk_forward.WALK_FORWARD_COLUMNS,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "walk_forward.xlsx"
            output = walk_forward.export_walk_forward_excel(
                detail,
                stock_id="2330",
                period="1y",
                strategy="score",
                train_days=20,
                test_days=10,
                step_days=10,
                output=str(path),
            )
            workbook = load_workbook(output)

        self.assertIn("Summary", workbook.sheetnames)
        self.assertIn("Detail", workbook.sheetnames)
        self.assertIn("Errors", workbook.sheetnames)

    def test_unsupported_sort_by_raises_value_error(self) -> None:
        with self.assertRaises(ValueError):
            walk_forward.run_walk_forward(
                "2330",
                sort_by="Test Sharpe Ratio",
                train_days=20,
                test_days=10,
            )

    def test_invalid_position_size_raises_value_error(self) -> None:
        with self.assertRaises(ValueError):
            walk_forward.run_walk_forward(
                "2330",
                position_size=0,
                train_days=20,
                test_days=10,
            )


if __name__ == "__main__":
    unittest.main()
